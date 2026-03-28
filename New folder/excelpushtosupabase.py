import os
import re
import requests
import gdown
import cloudinary
import cloudinary.uploader
from psycopg2 import pool
import psycopg2
from supabase import create_client, Client
import urllib.parse
import openpyxl # Thư viện đọc Excel có hyperlink
import sys
import json
import mimetypes
import unicodedata
sys.stdout.reconfigure(encoding='utf-8')

# ==========================================
# 1. CẤU HÌNH CLOUDINARY (Cho Video)
# ==========================================
cloudinary.config( 
    cloud_name = "dc8iezhmo", 
    api_key = "752322682664743", 
    api_secret = "cHpeGv8ipH2MGjM8lczcB3BgBPA" 
)

# ==========================================
# 2. CẤU HÌNH SUPABASE STORAGE & DATABASE
# ==========================================
url="https://jwkljthwnoxwnlgmcgdb.supabase.co"
key="eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imp3a2xqdGh3bm94d25sZ21jZ2RiIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MTMxMTkzMCwiZXhwIjoyMDg2ODg3OTMwfQ.m0hOXuRwn7hn1uXMb1ozpUuSUqpS3s0CyNCkpzdaM6o"
BUCKET_NAME = "mediaa" # Tên bucket bạn đã tạo trên Supabase
supabase = create_client(url, key)

# Cấu hình Connection Pool cho PostgreSQL (Supabase)
DB_HOST = "aws-1-ap-southeast-2.pooler.supabase.com" # Host của Supabase DB
DB_PORT = "6543" # Thường dùng port 6543 cho connection pooler (IPv4) hoặc 5432
DB_NAME = "postgres"
DB_USER = "postgres.jwkljthwnoxwnlgmcgdb"
DB_PASSWORD = "Saolaithe1@"

try:
    db_pool = pool.SimpleConnectionPool(1, 20, user=DB_USER, password=DB_PASSWORD, host=DB_HOST, port=DB_PORT, database=DB_NAME)
    if db_pool: print("✅ Kết nối Database thành công!")
except Exception as e:
    print(f"❌ Lỗi kết nối DB: {e}")
    exit(1)

# ==========================================
# 2. CÁC HÀM HỖ TRỢ XỬ LÝ MEDIA
# ==========================================
def get_drive_id(url):
    """Trích xuất ID từ link Drive"""
    if not url or "drive.google.com" not in str(url): return None
    try:
        if "/d/" in url: return url.split("/d/")[1].split("/")[0]
        if "id=" in url: return url.split("id=")[1].split("&")[0]
    except: return None
    return None

def sanitize_filename(text):
    """Chuyển tiếng Việt có dấu thành không dấu và thay khoảng trắng thành gạch dưới"""
    if not text:
        return "unknown"
    # Chuyển thành không dấu
    text = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode("utf-8")
    # Xóa các ký tự đặc biệt, chỉ giữ lại chữ, số và khoảng trắng
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    # Thay khoảng trắng và dấu gạch ngang thành dấu gạch dưới
    text = re.sub(r'[-\s]+', '_', text)
    return text

def get_extension(content_type):
    """Lấy đuôi file chuẩn từ Content-Type"""
    if content_type == 'image/jpeg': return 'jpg'
    if content_type == 'image/png': return 'png'
    if content_type == 'image/webp': return 'webp'
    if content_type == 'image/gif': return 'gif'
    if content_type == 'audio/mpeg': return 'mp3'
    if content_type == 'video/mp4': return 'mp4'
    ext = mimetypes.guess_extension(content_type)
    return ext.replace('.', '') if ext else 'jpg'

def upload_media_to_supabase(url, file_prefix, folder):
    """Tải file từ Drive/URL vào RAM và đẩy thẳng lên Supabase"""
    if not url or str(url).lower() in ['nan', 'none', '']: return None
    url = str(url).strip()
    
    # 1. Nếu là file local có sẵn trên máy
    if os.path.exists(url):
        content_type = mimetypes.guess_type(url)[0] or 'application/octet-stream'
        ext = get_extension(content_type)
        clean_prefix = sanitize_filename(file_prefix)
        storage_path = f"{folder}/{clean_prefix}.{ext}"
        try:
            with open(url, 'rb') as f:
                supabase.storage.from_(BUCKET_NAME).upload(
                    path=storage_path, file=f.read(),
                    file_options={"content-type": content_type, "upsert": "true"}
                )
            return supabase.storage.from_(BUCKET_NAME).get_public_url(storage_path)
        except Exception as e:
            print(f"   ⚠️ Lỗi upload file cục bộ: {e}")
            return None

    # 2. Nếu là link Drive hoặc Link mạng
    file_id = get_drive_id(url)
    download_url = f'https://drive.google.com/uc?export=download&id={file_id}' if file_id else url
    
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        res = requests.get(download_url, headers=headers, stream=True, timeout=15)
        
        if res.status_code == 200:
            content_type = res.headers.get('Content-Type', '')
            if 'text/html' in content_type:
                print(f"   ⚠️ Lỗi: Link bị Private! Vui lòng mở quyền truy cập.")
                return None
            
            ext = get_extension(content_type)
            clean_prefix = sanitize_filename(file_prefix)
            storage_path = f"{folder}/{clean_prefix}.{ext}"
            
            # Đẩy thẳng mảng byte (res.content) lên Supabase
            supabase.storage.from_(BUCKET_NAME).upload(
                path=storage_path, file=res.content,
                file_options={"content-type": content_type, "upsert": "true"}
            )
            return supabase.storage.from_(BUCKET_NAME).get_public_url(storage_path)
        else:
            print(f"   ❌ Lỗi tải {folder} (Status {res.status_code})")
            return None
    except Exception as e:
        print(f"   ⚠️ Lỗi mạng khi upload {folder}: {e}")
        return None

def upload_video_to_cloudinary(url):
    """Đẩy Video lên Cloudinary (hỗ trợ đọc thẳng URL hoặc truyền byte)"""
    if not url or str(url).lower() in ['nan', 'none', '']: return None
    url = str(url).strip()
    
    try:
        file_id = get_drive_id(url)
        if file_id:
            # Nếu là Drive, Cloudinary ko nhận direct link, ta tải về RAM rồi đẩy
            download_url = f'https://drive.google.com/uc?export=download&id={file_id}'
            res = requests.get(download_url, stream=True)
            if res.status_code == 200 and 'text/html' not in res.headers.get('Content-Type', ''):
                upload_res = cloudinary.uploader.upload(res.content, resource_type="video", folder="tapdoc_videos")
                return upload_res.get('secure_url')
            return None
        else:
            # Cloudinary hỗ trợ tự kéo từ URL thông thường hoặc Local Path
            upload_res = cloudinary.uploader.upload(url, resource_type="video", folder="tapdoc_videos")
            return upload_res.get('secure_url')
    except Exception as e:
        print(f"   ⚠️ Lỗi upload Video Cloudinary: {e}")
        return None

def get_str(val):
    return str(val).strip() if val is not None else ""

# ==========================================
# 3. HÀM CHÍNH - DUYỆT DATA VÀ LƯU TỪ VỰNG
# ==========================================
def process_data(xlsx_file_path):
    wb = openpyxl.load_workbook(xlsx_file_path, data_only=True)
    ws = wb.active
    headers = [get_str(cell.value) for cell in ws[2]]
    
    # 1. Câu lệnh Insert vào bảng vocabularies (trả về ID)
    insert_vocab_query = """
        INSERT INTO vocabularies (
            word, standard_meaning, simplified_meaning, 
            image_url, audio_url, video_url, is_required
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """
    
    # 2. Câu lệnh Insert vào bảng trung gian lesson_vocabularies
    insert_relation_query = """
        INSERT INTO lesson_vocabularies (lesson_id, vocab_id)
        VALUES (%s, %s)
        ON CONFLICT DO NOTHING
    """
    
    current_lesson_id = None
    
    for index, row in enumerate(ws.iter_rows(min_row=5), start=5):
        if row[0].value and (str(row[0].value).startswith("CHỦ ĐỀ") or str(row[0].value).startswith("TUẦN")):
            continue
            
        row_data = {}
        for idx, cell in enumerate(row):
            if idx < len(headers) and headers[idx]:
                col_name = headers[idx]
                val = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
                row_data[col_name] = val
        
        word_col_val = get_str(row_data.get('TỪ'))
        if not word_col_val or word_col_val.lower() in ['nan', 'none', '']:
            continue

        # XỬ LÝ LESSON_ID
        first_col_val = get_str(row[0].value)
        if first_col_val.lower().startswith("bài"):
            lesson_title = first_col_val.strip()
            try:
                conn = db_pool.getconn()
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM lessons WHERE title = %s", (lesson_title,))
                res = cursor.fetchone()
                current_lesson_id = res[0] if res else None
                cursor.close()
            finally:
                if conn: db_pool.putconn(conn)

        if not current_lesson_id:
            continue

        # CHUẨN BỊ DỮ LIỆU TỪ VỰNG
        raw_word = word_col_val
        is_required = "*" in raw_word or "*" in get_str(row_data.get('GHI CHÚ'))
        word = raw_word.replace("*", "").strip() 
        
        standard_meaning = get_str(row_data.get('GIẢI NGHĨA'))
        simplified_meaning = get_str(row_data.get('NGHĨA TINH GỌN'))
        
        # XỬ LÝ MEDIA
        img_val = get_str(row_data.get('HÌNH ẢNH')).split('\n')[0]
        vid_val = get_str(row_data.get('VIDEO MINH HỌA')).split('\n')[0]

        img_url_db = upload_media_to_supabase(img_val, f"{word}_image", "images")
        video_url_db = upload_video_to_cloudinary(vid_val)

        # THỰC THI LƯU DATABASE (2 BƯỚC)
        try:
            conn = db_pool.getconn()
            cursor = conn.cursor()
            
            # Bước 1: Insert vào vocabularies
            cursor.execute(insert_vocab_query, (
                word, standard_meaning, simplified_meaning, 
                img_url_db, None, video_url_db, is_required
            ))
            new_vocab_id = cursor.fetchone()[0]
            
            # Bước 2: Liên kết với lesson
            cursor.execute(insert_relation_query, (current_lesson_id, new_vocab_id))
            
            conn.commit()
            cursor.close()
            print(f"✅ Đã lưu '{word}' (ID: {new_vocab_id}) vào bài học {current_lesson_id}")
        except Exception as e:
            print(f"❌ Lỗi khi chèn DB '{word}': {e}")
            if conn: conn.rollback()
        finally:
            if conn: db_pool.putconn(conn)

if __name__ == "__main__":
    excel_file_path = "TẬP ĐỌC KNTT 2.2.xlsx"
    process_data(excel_file_path)
    if db_pool:
        db_pool.closeall()
        print("\n🏁 Đã đóng tất cả kết nối DB.")